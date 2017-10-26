from scipy.interpolate import *
from numpy import polyfit, polyval
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import openpyxl
from statistics import mean
import statistics
from openpyxl.drawing.image import Image

from matplotlib.font_manager import FontProperties
ChineseFont1 = FontProperties(fname = './msyh.ttf')

#改圖片範例程式碼在fill_A_table()最後幾行
#args
filename = 'input.xlsx'
output_name= 'output.xlsx'
polyfit_order=1
dot_size=2
cp_name_size=5

def get_fit_y(train_x, train_y, x):
	l = len(train_y)
	train_x = [ train_x[i] for i in range(l) if train_y[i]>0]
	train_y = [ train_y[i] for i in range(l) if train_y[i]>0]
	fit_func = polyfit( train_x, np.log(train_y), polyfit_order)
	y = np.exp( polyval( fit_func, x) )
	return y

def get_column(s, n):
	return [ e for e in zip(*s) ][n]

def find_set_by_l_grade(s, l_grade, func):
	func = func.upper()
	
	if func == 'ALL':
		return [e for e in s if e[0] == l_grade]

	return [e for e in s if e[0] == l_grade and e[2] == func]

def find_set(s, g_grade, func):
	func = func.upper()
	
	if func == 'ALL':
		return [e for e in s if e[1] == g_grade]

	return [e for e in s if e[1] == g_grade and e[2] == func]

def find_mid(s, g_grade, func):
	s = find_set(s, g_grade, func)
	if len(s) == 0:
		return -1
	t = get_column(s,4)
	return statistics.median(t)

def get_io(wb):
	io_ws = wb['io']
	
	F = io_ws['F'][1:]
	G = io_ws['G'][1:]
	L = io_ws['L'][1:]
	M = io_ws['M'][1:]
	P = io_ws['P'][1:]
	D = io_ws['D'][1:]
	R = io_ws['R'][1:]
	J = io_ws['J'][1:]
	K = io_ws['K'][1:]

	l_grade_list = [ f.value for f in F ]
	g_grade_list = [ g.value for g in G ]
	func_list = [l.value for l in L]
	cp_list = [cp.value for cp in M]
	salary_list = [s.value for s in P]
	name_list = [d.value for d in D]
	cr_list = [cr.value for cr in R]
	year_list = [y.value for y in J]
	ser_year_list = [s.value for s in K]

	s = []
	for e in zip(l_grade_list, g_grade_list, func_list, cp_list, salary_list, name_list, cr_list, year_list, ser_year_list):
		s.append(e)

	return s

def get_pr_salary(wb, g_grade, func):
	sheet_names = wb.get_sheet_names()
	for name in sheet_names:
		if func in name:
			ws = wb[name]
			salary_list = [ \
				ws.cell( \
				row =r , column=g_grade-1).value\
				for r in range(28,40)
				]
			# print(salary_list)
			return(salary_list)

def find_inter_pr(wb, g_grade, func, x):
	X_list = get_pr_salary(wb, g_grade, func)
	PR_list = [90,80,75,70,65,60,50,40,35,30,25,10]

	if x > X_list[0]:
		return PR_list[0]
	if x < X_list[-1]:
		return PR_list[-1]

	fit_func = interp1d( X_list, PR_list)
	y = fit_func(x)

	return round( min( max( 10, float(y) ), 90), 0 )

def check_A_correct(ws):
	g_grades = list(range(8,15))
	for g_grade in g_grades:
		tmp_list = [ ws.cell(row=r, column=g_grade-1).value for r in range(28,40) ]
		sort_list = tmp_list[:]
		sort_list.sort(reverse=True)
		if tmp_list != sort_list:
			# print(tmp_list)
			return False
	return True

def fill_A_table(wb, s):
	sheet_names = wb.get_sheet_names()
	g_grades = list(range(8,15))
	func_names = ['SALE+PM', 'RD', 'ALL']
	plt.rc('grid', linestyle=":", color=(0.7, 0.7, 0.7, 0.9))
	for name in sheet_names:
		ws = wb[name]
		for func in func_names:
			if func in name:
				# draw picture
				fig = plt.figure()
				ax = fig.add_subplot(111)
				ax.set_xlabel('salary(NTD)')
				ax.set_ylabel('grade')
				ax.yaxis.grid(True)

				if check_A_correct(ws):
					ax.set_title('A_{0}'.format(func))
				else:
					ax.set_title('A_{0} (table error)'.format(func))

				ax.get_yaxis().set_major_formatter(
    				matplotlib.ticker.FuncFormatter(lambda y, p: format(int(y), ',')))

				# fill table and plt cp_name
				for g_grade in g_grades:
					mid = find_mid(s, g_grade, func)
					pr = find_inter_pr(wb, g_grade, func, mid)
					ws.cell(row=26, column=g_grade-1).value = pr if mid >= 0 else 0
					ws.cell(row=27, column=g_grade-1).value = mid if mid >= 0 else 0

					# cp points
					tmp_list = find_set(s, g_grade, func)
					cp_x = [ cp[1] for cp in tmp_list if cp[3] == 'Y']
					cp_y = [ cp[4] for cp in tmp_list if cp[3] == 'Y']
					cp_names = [ cp[5] for cp in tmp_list if cp[3] == 'Y']

					# draw CP label for legend once
					if g_grades.index(g_grade) == 0:
						plt.scatter(cp_x, cp_y, color='black', marker='+', s=dot_size, label='CP')
					else:
						plt.scatter(cp_x, cp_y, color='black', marker='+', s=dot_size)
					# plot cp names string
					for xy in zip(cp_x, cp_y, cp_names):
					    ax.annotate('%s' % xy[2], xy=(xy[0], xy[1]),\
					     textcoords='data', fontproperties = ChineseFont1, size=cp_name_size)
				
				# plot lines
				# x = list(range(8,16))
				x = np.arange(8, 14, 0.1)

				rows = [28,30,34,38,39]
				labels = ['90', '75', '50', '25', '10']
				
				# AT-Med and trend
				y = [ ws.cell(row=27, column=g_grade-1).value for g_grade in g_grades ]
				dot_g_grades = [ g for g in g_grades if y[g_grades.index(g)] > 0 ]
				dot_y = [ e for e in y if e > 0 ]
				plt.scatter(dot_g_grades, dot_y, label='Med', color='red', s = dot_size)
				
				y = get_fit_y(g_grades, y, x)
				plt.plot(x, y, 'r', label='Med-Trend')

				# ggs trend
				for i in range(5):
					y = [ ws.cell(row=rows[i], column=g_grade-1).value for g_grade in g_grades ]
					y = get_fit_y(g_grades, y, x)
					if i == 2:
						plt.plot(x, y, 'g', label=labels[i])
					else:
						plt.plot(x, y, 'b')

				plt.legend()
				fig.tight_layout()
				fig.savefig( 'A_'+func+'.png', format = 'png', dpi = 300)

				# img = Image( 'A_'+func+'.png')
				# ws.add_image(img, 'S3')

def fill_io(wb, s):
	io_ws = wb['io']
	
	G = io_ws['G'][1:]
	L = io_ws['L'][1:]
	P = io_ws['P'][1:]

	for g in G:
		i = G.index(g)

		l = L[i]
		salary = P[i]

		io_ws.cell( row=i+2, column=17, value=find_inter_pr( wb, g.value, l.value, salary.value) )
		io_ws.cell( row=i+2, column=18, value=round( salary.value/find_mid(s, g.value, l.value), 2) )

	wb.save(output_name)

def draw_B_pic(wb, s):
	l_grades = list(range(10,23,2))
	func_names = ['SALE+PM', 'ALL', 'RD']
	plt.rc('grid', linestyle=":", color=(0.7, 0.7, 0.7, 0.9))
	
	for func in func_names:
		fig = plt.figure()
		fig, ax = plt.subplots()
		plt.grid(True)
			
		ax.set_title('B title')
		ax.set_xlabel('salary(NTD)')
		ax.set_ylabel('grade')
		ax.get_xaxis().set_major_formatter(
    				matplotlib.ticker.FuncFormatter(lambda x, p: format(int(x), ',')))

		for l_grade in l_grades:
			tmp_list = find_set_by_l_grade(s, l_grade, func)
			salary_list = [ e[4] for e in tmp_list]
			if len(salary_list) == 0:
				continue
			else:
				a = min(salary_list)
				b = max(salary_list)

			mid = (a+b)/2
			l = max(2000, b-a)
			
			plt1=plt.broken_barh( [(mid-l/2, l)], (l_grade-0.5,1),facecolors='blue')
		
		for name in wb.get_sheet_names():
			if func in name:
				ws = wb[name]
				for c in range(7):
					v = ws.cell(row=34, column=(c+7)).value
					plt2=plt.broken_barh( [(v-500, 1000)], (l_grades[c]-0.5,1),facecolors='red')


		plt.legend(handles=[plt1, plt2], labels=['blue_string', 'red_string'], loc='upper left')
		fig.tight_layout()
		fig.savefig( 'B_'+func+'.png', format = 'png', dpi = 300)

def fill_c_table(wb, s):
	cols = [2, 13, 25]
	func_names = ['ALL', 'RD', 'SALE+PM']
	l_grades = list(range(10,23,2))

	for name in wb.get_sheet_names():
		if 'C.' in name:
			ws = wb[name]
			for func in func_names:
				fig = plt.figure()
				fig, ax = plt.subplots()
				ax.set_title('C_{0}'.format(func))
				ax.set_xlabel('salary(NTD)')
				ax.set_ylabel('grade')
				ax.get_yaxis().set_major_formatter(
    				matplotlib.ticker.FuncFormatter(lambda y, p: format(int(y), ',')))
				ax.yaxis.grid(True)

				col_bias = cols[ func_names.index(func) ]
				max_list = []
				med_list = []
				min_list = []

				for l_grade in l_grades:
					c = l_grades.index(l_grade) + col_bias
					max_v = med_v = min_v = 0

					tmp_list = find_set_by_l_grade(s, l_grade, func)
					salary_list = [ e[4] for e in tmp_list]

					l = len(salary_list)					
					if l > 0:
						max_v = max(salary_list) //1
						med_v = statistics.median(salary_list) //1
						min_v = min(salary_list) //1

					ws.cell(row=4, column=c).value = l
					ws.cell(row=5, column=c).value = max_v
					ws.cell(row=6, column=c).value = med_v
					ws.cell(row=7, column=c).value = min_v
					
					max_list.append(max_v)
					med_list.append(med_v)
					min_list.append(min_v)

					people_mark = plt.scatter([l_grade]*len(salary_list), salary_list,
						 s=dot_size,
						 color='black')

				smooth_l_grades = np.arange(10, 22, 0.1)
				y = get_fit_y(l_grades, max_list, smooth_l_grades)
				# fit_func = polyfit( l_grades, max_list, polyfit_order)
				# y = polyval( fit_func, smooth_l_grades)
				bb, = plt.plot(smooth_l_grades, y, label='max')
				y = get_fit_y(l_grades, med_list, smooth_l_grades)
				# fit_func = polyfit( l_grades, med_list, polyfit_order)
				# y = polyval( fit_func, smooth_l_grades)
				cc, = plt.plot(smooth_l_grades, y, label='med')
				y = get_fit_y(l_grades, min_list, smooth_l_grades)
				# fit_func = polyfit( l_grades, min_list, polyfit_order)
				# y = polyval( fit_func, smooth_l_grades)
				dd, = plt.plot(smooth_l_grades, y, label='min')

				# plt.legend()
				plt.legend(handles=[people_mark, bb,cc,dd], labels=['p','max', 'med', 'min'], loc='upper left')
				fig.tight_layout()
				fig.savefig( 'C_'+func+'.png', format = 'png', dpi = 300)

def l_to_g_grade(l_grade):
	l_grades = list(range(10, 23, 2))
	g_grades = list(range(8,15))
	return g_grades[l_grades.index(l_grade)]

def fill_d_table(wb, s):
	rows = [2, 13, 23]
	func_names = ['ALL', 'RD', 'SALE+PM']
	l_grades = list(range(22, 9,-2))

	for name in wb.get_sheet_names():
		if 'D.' in name:
			ws = wb[name]
			for func in func_names:
				r_bias = rows[ func_names.index(func) ]
				for l_grade in l_grades:
					r = l_grades.index(l_grade) + r_bias
					g_grade = l_to_g_grade(l_grade)

					tmp_list = find_set_by_l_grade(s, l_grade, func)
					
					# print(tmp_list)
					l = len(tmp_list)
					if l >0:
						mean_year = mean([ e[7] for e in tmp_list])
						mean_ser_year = mean([ e[8] for e in tmp_list])

						med_salary = statistics.median([ e[4] for e in tmp_list])
						med_p_value = find_inter_pr(wb, g_grade,func, med_salary)
						min_salary = min([ e[4] for e in tmp_list])
						med_salary = statistics.median([ e[4] for e in tmp_list])
						max_salary = max([ e[4] for e in tmp_list])

						ws.cell(row=r, column=3).value = l
						ws.cell(row=r, column=5).value = round(mean_year,0)
						ws.cell(row=r, column=6).value = round(mean_ser_year,1)
						ws.cell(row=r, column=7).value = med_p_value
						ws.cell(row=r, column=8).number_format = '#,##0'
						ws.cell(row=r, column=8).value = min_salary
						ws.cell(row=r, column=9).number_format = '#,##0'
						ws.cell(row=r, column=9).value = med_salary
						ws.cell(row=r, column=10).number_format = '#,##0'
						ws.cell(row=r, column=10).value = max_salary

def draw_E_pic(wb):
	cols = [1, 12, 24]
	func_names = ['ALL', 'RD', 'SALE+PM']
	l_grades = list(range(10,23,2))
	plt.gca()

	for name in wb.get_sheet_names():
		if 'E.' in name:
			ws = wb[name]
			for func in func_names:
				# fig = plt.figure()
				fig, ax = plt.subplots()
				ax.set_title('E_{0}'.format(func))
				ax.set_xlabel('salary(NTD)')
				ax.set_ylabel('grade')
				ax.get_yaxis().set_major_formatter(
    				matplotlib.ticker.FuncFormatter(lambda y, p: format(int(y), ',')))
				ax.yaxis.grid(True)

				col_bias = cols[ func_names.index(func) ]
				max_list = [ e.value for e in ws[7][col_bias:col_bias+7] ]
				med_list = [ e.value for e in ws[8][col_bias:col_bias+7] ]
				min_list = [ e.value for e in ws[9][col_bias:col_bias+7] ]
				
				# draw scatter
				for l_grade in l_grades:
					tmp_list = find_set_by_l_grade(s, l_grade, func)
					salary_list = [ e[4] for e in tmp_list]
					people_mark = plt.scatter([l_grade]*len(salary_list), salary_list,
						 s=2,
						 color='black')

				if None in max_list:
					return

				smooth_l_grades = np.arange(10, 22, 0.1)
				y = get_fit_y(l_grades, max_list, smooth_l_grades)
				bb, = plt.plot(smooth_l_grades, y, label='max')
				y = get_fit_y(l_grades, med_list, smooth_l_grades)
				cc, = plt.plot(smooth_l_grades, y, label='med')
				y = get_fit_y(l_grades, min_list, smooth_l_grades)
				dd, = plt.plot(smooth_l_grades, y, label='min')

				plt.legend(handles=[people_mark, bb,cc,dd], labels=['p','max', 'med', 'min'], loc='upper left')
				fig.tight_layout()
				fig.savefig( 'E_'+func+'.png', format = 'png', dpi = 300)		


wb = load_workbook(filename = filename)

s = get_io(wb)
fill_A_table(wb, s)
fill_io(wb, s)

# get new io data
s = get_io(wb)
draw_B_pic(wb,s)
fill_c_table(wb,s)
fill_d_table(wb,s)
draw_E_pic(wb)

wb.save(filename)
